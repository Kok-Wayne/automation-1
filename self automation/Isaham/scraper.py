"""
Bursa Malaysia Contact Info Scraper
=====================================
Source: https://www.isaham.my/all-sectors?style=new
Extracts: Phone, Email, Address from each listed company's website
Output: blue_chip.xlsx

Requirements:
    pip install selenium webdriver-manager requests beautifulsoup4 openpyxl lxml

Usage:
    python scraper.py
    python scraper.py --sector semiconductors --max 20
"""

import re
import os
import json
import time
import random
import argparse
import logging
from dataclasses import dataclass, field, asdict
from typing import Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

from contact_extracted import scrape_company_contacts_v2

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Data model ────────────────────────────────────────────────────────────────
@dataclass
class CompanyInfo:
    stock_code:    str = ""
    company_name:  str = ""
    sector:        str = ""
    isaham_url:    str = ""
    company_url:   str = ""
    phone:         str = ""
    email:         str = ""
    address:       str = ""
    scrape_status: str = ""

# ── Regex patterns ─────────────────────────────────────────────────────────────
PHONE_RE = re.compile(
    r'(?:\+?60[-.\s]?|0)'          # country code or leading 0
    r'(?:\d[-.\s]?){7,11}',        # 7-11 digits with optional separators
)
EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')

CONTACT_KEYWORDS = [
    "contact", "contact-us", "contacts", "about", "about-us",
    "hubungi", "kenalan", "info",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Selenium helpers ──────────────────────────────────────────────────────────
def build_driver(headless: bool = True) -> webdriver.Chrome:
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(f"user-agent={HEADERS['User-Agent']}")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"},
    )
    return driver


def get_sector_stocks(driver: webdriver.Chrome, sector_slug: str) -> list[dict]:
    """
    Visit a sector page on isaham.my and return list of {code, name, url}.
    The stock table is rendered by JS, so we need Selenium.
    """
    url = f"https://www.isaham.my/sector/{sector_slug}"
    log.info(f"  Opening sector page: {url}")
    driver.get(url)
    time.sleep(3)  # let JS render

    stocks = []
    try:
        # Wait for the stock table rows
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr, .stock-row, .sector-stock"))
        )
    except TimeoutException:
        log.warning(f"  Timed out waiting for stocks on {url}")

    # Try common table selectors isaham uses
    rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")
    if not rows:
        rows = driver.find_elements(By.CSS_SELECTOR, ".stock-list-item, .sector-item")

    for row in rows:
        try:
            links = row.find_elements(By.TAG_NAME, "a")
            for link in links:
                href = link.get_attribute("href") or ""
                text = link.text.strip()
                # isaham stock pages look like /stock/XXXX
                if "/stock/" in href:
                    parts = href.rstrip("/").split("/")
                    code = parts[-1].upper()
                    name = text if text else code
                    stocks.append({
                        "code": code,
                        "name": name,
                        "isaham_url": href,
                    })
                    break
        except Exception:
            continue

    # Deduplicate by code
    seen = set()
    unique = []
    for s in stocks:
        if s["code"] not in seen:
            seen.add(s["code"])
            unique.append(s)

    log.info(f"  Found {len(unique)} stocks in sector")
    return unique


def get_company_website(driver: webdriver.Chrome, isaham_url: str, retries: int = 2) -> str:
    """Visit isaham stock page and find the company's official website URL."""
    log.info(f"    Fetching stock page: {isaham_url}")

    for attempt in range(retries + 1):
        try:
            driver.set_page_load_timeout(20)
            driver.get(isaham_url)
            break
        except TimeoutException:
            if attempt < retries:
                log.warning(f"    Timeout loading {isaham_url}, retry {attempt + 1}/{retries} …")
                try:
                    driver.execute_script("window.stop();")
                except Exception:
                    pass
                time.sleep(2)
            else:
                log.warning(f"    Gave up loading {isaham_url} after {retries} retries")
                return ""
        except Exception as e:
            log.warning(f"    Error loading {isaham_url}: {e}")
            return ""
        finally:
            driver.set_page_load_timeout(30)

    time.sleep(2)

    # isaham shows the official website in a detail panel; look for external links
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a[href]"))
        )
    except TimeoutException:
        pass

    for a in driver.find_elements(By.TAG_NAME, "a"):
        href = a.get_attribute("href") or ""
        # Skip internal / social / common non-company links
        if any(skip in href for skip in [
            "isaham", "facebook", "twitter", "instagram", "linkedin",
            "bursamalaysia", "klse", "google", "youtube", "t.me", "mailto",
            "javascript", "#",
        ]):
            continue
        parsed = urlparse(href)
        if parsed.scheme in ("http", "https") and parsed.netloc:
            return href.rstrip("/")

    return ""


# ── Contact scraping helpers ──────────────────────────────────────────────────
def clean_phone(raw: str) -> str:
    """Normalise phone: strip spaces/dashes, keep leading +."""
    digits = re.sub(r"[^\d+]", "", raw)
    return digits if len(digits) >= 8 else ""


def extract_contacts_from_html(html: str, base_url: str) -> dict:
    """Parse raw HTML and return {phone, email, address}."""
    soup = BeautifulSoup(html, "lxml")

    # Remove script/style noise
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    text = soup.get_text(separator=" ", strip=True)

    # Phone
    phones = [clean_phone(m) for m in PHONE_RE.findall(text)]
    phones = [p for p in phones if p]
    phone = phones[0] if phones else ""

    # Email — exclude image/icon file extensions and common placeholder domains
    emails = EMAIL_RE.findall(text)
    emails = [
        e for e in emails
        if not re.search(r'\.(png|jpg|gif|svg|webp|ico|css|js)$', e, re.I)
        and "example" not in e and "yourdomain" not in e
    ]
    email = emails[0] if emails else ""

    # Address — look for structured address tags first
    address = ""
    for tag in soup.find_all(["address", "p", "div", "span"]):
        t = tag.get_text(separator=" ", strip=True)
        # Heuristic: contains a postcode-like pattern (5 digits) or "Jalan/Lot/No."
        if re.search(r'\b\d{5}\b', t) or re.search(r'\b(Jalan|Jln|Lot|No\.|Level|Floor|Suite|Block)\b', t, re.I):
            address = re.sub(r'\s+', ' ', t)[:300]
            break

    return {"phone": phone, "email": email, "address": address}


def find_contact_page(session: requests.Session, base_url: str) -> Optional[str]:
    """Try to find a /contact-us style page; return its URL or None."""
    parsed = urlparse(base_url)
    root = f"{parsed.scheme}://{parsed.netloc}"

    for kw in CONTACT_KEYWORDS:
        for sep in ("/", "-", "_"):
            candidate = f"{root}/{kw.replace('-', sep)}"
            try:
                r = session.get(candidate, timeout=8, allow_redirects=True)
                if r.status_code == 200 and len(r.text) > 500:
                    return candidate
            except Exception:
                continue
    return None


def scrape_company_contacts(website_url: str) -> dict:
    """
    Visit the company website (homepage + contact page) and
    return the best phone/email/address found.
    """
    session = requests.Session()
    session.headers.update(HEADERS)
    result = {"phone": "", "email": "", "address": ""}

    pages_to_try = [website_url]
    contact_page = find_contact_page(session, website_url)
    if contact_page:
        pages_to_try.append(contact_page)

    for url in pages_to_try:
        try:
            r = session.get(url, timeout=12, allow_redirects=True)
            r.raise_for_status()
            data = extract_contacts_from_html(r.text, url)
            # Fill in missing fields
            for key in ("phone", "email", "address"):
                if not result[key] and data[key]:
                    result[key] = data[key]
        except Exception as e:
            log.warning(f"    Could not fetch {url}: {e}")

        # Stop early if we have all three
        if all(result.values()):
            break

        time.sleep(random.uniform(0.5, 1.5))

    return result


# ── Excel export ──────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
ALT_FILL     = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT    = Font(name="Arial", size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN         = Side(style="thin", color="B0C4DE")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Stock Code",    12),
    ("Company Name",  30),
    ("Sector",        20),
    ("Website",       35),
    ("Phone",         18),
    ("Email",         30),
    ("Address",       50),
    ("Status",        14),
]


def export_to_excel(records: list[CompanyInfo], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Contacts"
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font    = HEADER_FONT
        cell.fill    = HEADER_FILL
        cell.alignment = CENTER
        cell.border  = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Data rows
    fields = ["stock_code", "company_name", "sector", "company_url",
              "phone", "email", "address", "scrape_status"]

    for row_idx, rec in enumerate(records, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        d = asdict(rec)
        for col_idx, key in enumerate(fields, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=d.get(key, ""))
            cell.font    = CELL_FONT
            cell.border  = BORDER
            cell.alignment = CENTER if col_idx in (1, 8) else LEFT
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 18

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    from pathlib import Path
    save_path = output_path
    counter = 1
    while True:
        try:
            wb.save(save_path)
            if save_path != output_path:
                log.info(f"  (原文件被占用，已另存为 {save_path})")
            break
        except PermissionError:
            stem = Path(output_path).stem
            suffix = Path(output_path).suffix
            save_path = str(Path(output_path).parent / f"{stem}_{counter}{suffix}")
            counter += 1

# ── 断点续跑 (resume / checkpoint) ──────────────────────────────────────────
def load_progress(progress_file: str) -> dict:
    """
    读取上次运行留下的进度文件。文件不存在或损坏时返回一个空的初始状态。
    结构:
    {
        "completed_sectors": ["sector-slug-1", "sector-slug-2", ...],  # 已经完整跑完的 sector
        "current_sector_slug": "semiconductors",                       # 正在进行中、还没跑完的 sector
        "current_sector_done_codes": ["1234", "5678"]                  # 该 sector 内已经抓过的股票代码
    }
    """
    if not os.path.exists(progress_file):
        return {"completed_sectors": [], "current_sector_slug": "", "current_sector_done_codes": []}
    try:
        with open(progress_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        data.setdefault("completed_sectors", [])
        data.setdefault("current_sector_slug", "")
        data.setdefault("current_sector_done_codes", [])
        return data
    except Exception as e:
        log.warning(f"进度文件读取失败 ({e})，将从头开始")
        return {"completed_sectors": [], "current_sector_slug": "", "current_sector_done_codes": []}


def save_progress(progress_file: str, completed_sectors: set, current_sector_slug: str, current_sector_done_codes: set):
    """把当前进度写入磁盘（每抓完一家公司调用一次，所以随时可以安全中断）。"""
    data = {
        "completed_sectors": sorted(completed_sectors),
        "current_sector_slug": current_sector_slug,
        "current_sector_done_codes": sorted(current_sector_done_codes),
    }
    tmp_file = progress_file + ".tmp"
    try:
        with open(tmp_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp_file, progress_file)  # 原子替换，避免写入过程中被中断导致文件损坏
    except Exception as e:
        log.warning(f"进度保存失败: {e}")


# ── Main orchestration ────────────────────────────────────────────────────────
def get_all_sector_slugs(driver: webdriver.Chrome) -> list[tuple[str, str]]:
    """Return list of (sector_name, sector_slug) from the All Sectors page."""
    log.info("Loading all-sectors page …")
    driver.get("https://www.isaham.my/all-sectors")
    time.sleep(3)

    # Wait for sector links with the correct class
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "a.list-sector-item"))
        )
    except TimeoutException:
        log.warning("Timeout waiting for sector list, trying anyway …")

    slugs = []
    seen  = set()
    for a in driver.find_elements(By.CSS_SELECTOR, "a.list-sector-item"):
        href = a.get_attribute("href") or ""
        if "/sector/" not in href:
            continue
        slug = href.rstrip("/").split("/sector/")[-1]
        name = a.text.strip().split("\n")[0] or slug
        if slug and slug not in seen:
            seen.add(slug)
            slugs.append((name, slug))

    log.info(f"Found {len(slugs)} sectors")
    return slugs


def safe_filename(name: str) -> str:
    """Convert sector name to a safe filename."""
    name = re.sub(r'[\\/*?:"<>|]', "", name)
    name = name.strip().replace(" ", "_")
    return name or "unknown_sector"


class SectorWriter:
    """
    Keeps one Excel workbook open per sector.
    Call .append(rec) after each company — writes immediately.
    Call .close() when the sector is done.
    """
    FIELDS = ["stock_code", "company_name", "sector", "company_url",
              "phone", "email", "address", "scrape_status"]

    def __init__(self, sector_name: str, resuming: bool = False):
        self.sector_name = sector_name
        self.path = f"{safe_filename(sector_name)}.xlsx"

        if resuming and os.path.exists(self.path):
            # 断点续跑：打开已有文件，从最后一行之后继续写，而不是覆盖重建
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active
            self.row_idx = self.ws.max_row + 1
            log.info(f"  📂 继续使用已有文件 {self.path}（从第 {self.row_idx} 行续写）")
        else:
            self.row_idx = 2
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = sector_name[:31]  # Excel sheet name max 31 chars
            self.ws.freeze_panes = "A2"
            self._write_header()
            self._save()
            log.info(f"  📄 Created {self.path}")

    def _write_header(self):
        for col_idx, (header, width) in enumerate(COLUMNS, start=1):
            cell = self.ws.cell(row=1, column=col_idx, value=header)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = BORDER
            self.ws.column_dimensions[get_column_letter(col_idx)].width = width
        self.ws.row_dimensions[1].height = 22

    def append(self, rec: "CompanyInfo"):
        d = asdict(rec)
        fill = ALT_FILL if self.row_idx % 2 == 0 else None
        for col_idx, key in enumerate(self.FIELDS, start=1):
            cell = self.ws.cell(row=self.row_idx, column=col_idx, value=d.get(key, ""))
            cell.font      = CELL_FONT
            cell.border    = BORDER
            cell.alignment = CENTER if col_idx in (1, 8) else LEFT
            if fill:
                cell.fill = fill
        self.ws.row_dimensions[self.row_idx].height = 18
        self.row_idx += 1
        self._save()
        log.info(f"    ✓ Written to {self.path}  (row {self.row_idx - 1})")

    def _save(self):
        from pathlib import Path
        save_path = self.path
        counter = 1
        while True:
            try:
                self.wb.save(save_path)
                break
            except PermissionError:
                stem   = Path(self.path).stem
                suffix = Path(self.path).suffix
                save_path = str(Path(self.path).parent / f"{stem}_{counter}{suffix}")
                counter += 1

    def close(self):
        # Apply auto-filter on final save
        self.ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
        self._save()
        log.info(f"  💾 Sector done → {self.path}  ({self.row_idx - 2} companies)")


def run(sector_filter: str = "", max_companies: int = 0, headless: bool = True, output: str = "blue_chip.xlsx",
        progress_file: str = "scrape_progress.json", reset: bool = False):

    if reset and os.path.exists(progress_file):
        os.remove(progress_file)
        log.info("🔄 已重置进度文件，将从头开始抓取")

    progress = load_progress(progress_file)
    completed_sectors = set(progress["completed_sectors"])
    current_sector_slug = progress["current_sector_slug"]
    current_sector_done_codes = set(progress["current_sector_done_codes"])

    if completed_sectors or current_sector_done_codes:
        log.info(f"📌 检测到上次的进度：已完成 {len(completed_sectors)} 个 sector"
                 + (f"，当前 sector「{current_sector_slug}」已抓 {len(current_sector_done_codes)} 家公司" if current_sector_slug else ""))
        log.info("   将自动跳过已完成的部分，从断点继续。如需从头开始，请加 --reset 参数")

    driver = build_driver(headless=headless)
    all_records: list[CompanyInfo] = []

    try:
        if sector_filter:
            sectors = [(sector_filter.replace("-", " ").title(), sector_filter)]
        else:
            sectors = get_all_sector_slugs(driver)

        for sector_name, sector_slug in sectors:
            if sector_slug in completed_sectors:
                log.info(f"\n⏭  跳过已完成的 sector: {sector_name} ({sector_slug})")
                continue

            log.info(f"\n── Sector: {sector_name} ({sector_slug}) ──")
            stocks = get_sector_stocks(driver, sector_slug)

            if max_companies:
                stocks = stocks[:max_companies]

            # 如果这个 sector 正是上次中断时进行中的那个，沿用已完成的代码集合；否则从空集合开始
            if sector_slug == current_sector_slug:
                done_codes = set(current_sector_done_codes)
            else:
                done_codes = set()

            writer = SectorWriter(sector_name, resuming=bool(done_codes))

            for stock in stocks:
                if stock["code"] in done_codes:
                    log.info(f"  ⏭ {stock['code']} — {stock['name']}（已抓过，跳过）")
                    continue

                rec = CompanyInfo(
                    stock_code=stock["code"],
                    company_name=stock["name"],
                    sector=sector_name,
                    isaham_url=stock["isaham_url"],
                )

                log.info(f"  ▶ {stock['code']} — {stock['name']}")

                # Step 1: get official website from isaham page
                website = get_company_website(driver, stock["isaham_url"])
                rec.company_url = website

                if not website:
                    rec.scrape_status = "No website found"
                    log.warning(f"    No website found")
                    writer.append(rec)   # write immediately
                    all_records.append(rec)
                else:
                    # Step 2: scrape contact info from official website
                    log.info(f"    Website: {website}")
                    contacts = scrape_company_contacts_v2(website, driver=driver)
                    rec.phone   = contacts["phone"]
                    rec.email   = contacts["email"]

                    found = [k for k, v in contacts.items() if v]
                    rec.scrape_status = "OK" if found else "No contacts found"
                    log.info(f"    → phone={rec.phone or '—'}  email={rec.email or '—'}")

                    writer.append(rec)       # write immediately
                    all_records.append(rec)
                    time.sleep(random.uniform(1, 2))

                # 这家公司抓完了（无论成功与否），更新断点
                done_codes.add(stock["code"])
                save_progress(progress_file, completed_sectors, sector_slug, done_codes)

            writer.close()

            # 整个 sector 跑完了，标记为已完成，清空"进行中"的状态
            completed_sectors.add(sector_slug)
            current_sector_slug = ""
            done_codes = set()
            save_progress(progress_file, completed_sectors, "", done_codes)

    finally:
        driver.quit()

    log.info(f"\n✅ Done. {len(all_records)} companies scraped across {len(sectors)} sectors")
    log.info(f"   （如果这次是完整跑完所有 sector，可以删除 {progress_file}；否则下次运行会自动从断点继续）")
    return all_records


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Bursa contact scraper")
    parser.add_argument("--sector", default="",      help="Scrape a single sector slug (e.g. semiconductors)")
    parser.add_argument("--max",    default=0,  type=int, help="Max companies per sector (0 = all)")
    parser.add_argument("--output", default="blue_chip.xlsx", help="Output Excel filename")
    parser.add_argument("--visible", action="store_true",            help="Show browser window (non-headless)")
    parser.add_argument("--progress-file", default="scrape_progress.json",
                         help="断点续跑的进度文件路径 (默认 scrape_progress.json)")
    parser.add_argument("--reset", action="store_true",
                         help="忽略并删除之前的进度，从头开始抓取")
    args = parser.parse_args()

    run(
        sector_filter=args.sector,
        max_companies=args.max,
        headless=not args.visible,
        output=args.output,
        progress_file=args.progress_file,
        reset=args.reset,
    )