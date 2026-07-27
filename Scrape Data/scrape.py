"""
Google Maps Scraper — core logic
Called by scraper_ui.py
"""

from playwright.sync_api import sync_playwright
import time, os, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def interruptible_sleep(seconds, stop_event=None):
    """Sleep that can be interrupted by stop_event every 0.2s."""
    steps = int(seconds / 0.2)
    for _ in range(steps):
        if stop_event and stop_event.is_set():
            raise InterruptedError("Stopped by user")
        time.sleep(0.2)


def safe_text(page, selector, timeout=2000):
    try:
        return page.locator(selector).first.inner_text(timeout=timeout).strip()
    except:
        return ""


def safe_attr(page, selector, attr, timeout=2000):
    try:
        return page.locator(selector).first.get_attribute(attr, timeout=timeout) or ""
    except:
        return ""


def extract_state(address):
    states = [
        "Johor", "Kedah", "Kelantan", "Melaka", "Negeri Sembilan",
        "Pahang", "Perak", "Perlis", "Penang", "Pulau Pinang",
        "Sabah", "Sarawak", "Selangor", "Terengganu",
        "Kuala Lumpur", "WP Kuala Lumpur", "Labuan", "Putrajaya"
    ]
    for s in states:
        if s.lower() in address.lower():
            return s
    return ""


def extract_place_data(page):
    name = safe_text(page, 'h1[class*="DUwDvf"]') or safe_text(page, 'h1')
    address = safe_text(page, 'button[data-item-id="address"] .Io6YTe') \
           or safe_text(page, '[data-item-id="address"]')
    tel = safe_text(page, 'button[data-item-id*="phone"] .Io6YTe') \
       or safe_text(page, '[data-tooltip="Copy phone number"]')
    website = safe_attr(page, 'a[data-item-id="authority"]', 'href') \
           or safe_text(page, 'a[data-item-id="authority"]')
    place_type = safe_text(page, 'button[jsaction*="category"]') \
               or safe_text(page, '.DkEaL')
    state = extract_state(address)

    # Skip if temporarily closed
    page_text = safe_text(page, 'div[class*="o0Svhf"]') \
             or safe_text(page, '.ZDu9vd') \
             or safe_text(page, '[aria-label*="losed"]')
    if any(kw in page_text.lower() for kw in ["temporarily closed", "sementara ditutup"]):
        print(f"    Skipped (temporarily closed): {name}")
        return None

    return {
        "Name": name, "Address": address, "Tel": tel,
        "State": state, "Website": website, "Type": place_type,
    } if name else None


def scroll_sidebar(page, times, wait, stop_event=None):
    try:
        feed = page.locator('div[role="feed"]')
        for i in range(times):
            if stop_event and stop_event.is_set():
                print("\n  Scroll stopped by user.")
                break
            feed.evaluate("el => el.scrollBy(0, 1200)")
            interruptible_sleep(wait, stop_event)
            print(f"  Scrolling... {i+1}/{times}", end="\r")
        print()
    except InterruptedError:
        print("\n  Stopped by user.")
    except Exception as e:
        print(f"  Scroll error: {e}")


def load_existing(full_path, fields):
    existing = set()
    if not os.path.exists(full_path):
        return existing
    wb = openpyxl.load_workbook(full_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    name_col = headers.index("Name") if "Name" in headers else 0
    tel_col  = headers.index("Tel")  if "Tel"  in headers else 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (row[name_col] or "").lower()
        tel  = row[tel_col]  or ""
        existing.add((name, tel))
    return existing


def save_results(results, full_path, fields):
    file_exists = os.path.exists(full_path)

    if file_exists:
        wb = openpyxl.load_workbook(full_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        header_fill = PatternFill("solid", fgColor="2E75B6")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        widths = {"Name": 35, "Follow Up": 15, "Doctor": 25,
                  "Doctor Tel": 15, "Tel": 18, "Address": 55,
                  "State": 18, "Website": 40, "Type": 25}
        for col, field in enumerate(fields, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = widths.get(field, 20)

    for record in results:
        ws.append([record.get(f, "") for f in fields])
        # Follow Up, Doctor, Doctor Tel are always empty (to be filled manually)

    wb.save(full_path)


def scrape(query, output_file, scroll_times, headless, stop_event=None, exclude_kws=None):
    results = []
    fields = ["Name", "Follow Up", "Doctor", "Doctor Tel", "Tel", "Address", "State", "Website", "Type"]
    full_path = os.path.join(SCRIPT_DIR, output_file)

    seen = load_existing(full_path, fields)
    existing_count = len(seen)
    if existing_count:
        print(f"\n  Found {existing_count} existing records in '{output_file}', new entries will be appended.")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        page = browser.new_page()

        url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
        print(f"\n  Opening: {url}")
        page.goto(url)
        interruptible_sleep(3, stop_event)

        scroll_sidebar(page, scroll_times, 1.5, stop_event)

        links = page.locator('a[href*="/maps/place/"]').all()
        print(f"  Found {len(links)} listings\n")

        for i, link in enumerate(links):
            try:
                if stop_event and stop_event.is_set():
                    print("  Stopped by user.")
                    break
                link.click()
                interruptible_sleep(2, stop_event)
                data = extract_place_data(page)
                if data:
                    # Check exclude keywords against name
                    name_lower = data["Name"].lower()
                    excluded = exclude_kws and any(kw in name_lower for kw in exclude_kws)
                    if excluded:
                        matched = next(kw for kw in exclude_kws if kw in name_lower)
                        print(f"  [{i+1}] Excluded ('{matched}'): {data['Name']}")
                    else:
                        key = (name_lower, data["Tel"])
                        if key in seen:
                            print(f"  [{i+1}] Duplicate, skipped: {data['Name']}")
                        else:
                            seen.add(key)
                            results.append(data)
                            print(f"  [{i+1}] {data['Name']} | {data['Tel']}")
                else:
                    print(f"  [{i+1}] Skipped (no name found)")
            except Exception as e:
                print(f"  [{i+1}] Error: {e}")

        browser.close()

    if not results:
        print("\n  No new data found.")
        return

    try:
        save_results(results, full_path, fields)
    except PermissionError:
        print(f"\n  ERROR: Cannot save '{full_path}'")
        print("  The file may be open in Excel — close it and try again.")
        print("  Saving to Desktop instead...")
        full_path = os.path.join(os.path.expanduser("~"), "Desktop", output_file)
        save_results(results, full_path, fields)

    print(f"\n  {len(results)} new entries appended -> {full_path}")
    print(f"  Total records in file: {existing_count + len(results)}\n")


EXCLUDE_CITIES = [
    "Johor Bahru", "Kuala Lumpur", "Petaling Jaya",
    "Shah Alam", "George Town", "Ipoh", "Kota Kinabalu",
    "Kuching", "Alor Setar", "Kota Bharu", "Kuantan",
]


def extract_area_from_address(address):
    if not address:
        return None
    parts = [p.strip() for p in address.split(",")]
    for part in parts:
        match = re.match(r'^\d{5}\s+(.+)$', part)
        if match:
            city = match.group(1).strip()
            if city not in EXCLUDE_CITIES:
                return city
    for part in parts:
        if any(kw in part for kw in ["Taman", "Bandar", "Pandan", "Damansara"]):
            return part.strip()
    return None


def discover_areas(base_query, city, quick_scrolls=20, stop_event=None):
    areas_found = set()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        url = f"https://www.google.com/maps/search/{base_query.replace(' ', '+')}"
        print(f"\n  Scanning: {url}")
        page.goto(url)
        interruptible_sleep(3, stop_event)

        scroll_sidebar(page, quick_scrolls, 1.5, stop_event)

        links = page.locator('a[href*="/maps/place/"]').all()
        print(f"  Found {len(links)} listings to scan for areas\n")

        for link in links:
            try:
                if stop_event and stop_event.is_set():
                    break
                link.click()
                interruptible_sleep(1.5, stop_event)
                address = safe_text(page, 'button[data-item-id="address"] .Io6YTe') \
                       or safe_text(page, '[data-item-id="address"]')
                area = extract_area_from_address(address)
                if area:
                    areas_found.add(area)
            except:
                continue

        browser.close()

    return sorted(list(areas_found))
